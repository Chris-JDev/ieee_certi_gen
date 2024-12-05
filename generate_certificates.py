import os
from openpyxl import load_workbook
from PIL import Image, ImageDraw, ImageFont

# File paths
excel_file = "Presenting Authors.xlsx"  # Your Excel file
template_image = "image.png"  # Certificate template image
output_folder = "Certificates"  # Where certificates will be saved

# Read data from Excel
def load_excel_data(file_path):
    workbook = load_workbook(file_path)
    sheet = workbook.active
    data = []
    for row in sheet.iter_rows(min_row=2, values_only=True):  # Skip the header row
        name, title = row
        if name and title:
            data.append((name, title))
    return data

# Helper function to wrap text
def wrap_text(text, font, max_width, draw):
    words = text.split()
    lines = []
    current_line = ""

    for word in words:
        test_line = current_line + " " + word if current_line else word
        # Use textbbox() to measure text width and height
        line_bbox = draw.textbbox((0, 0), test_line, font=font)
        line_width = line_bbox[2] - line_bbox[0]
        if line_width <= max_width:
            current_line = test_line
        else:
            lines.append(current_line)
            current_line = word
    if current_line:
        lines.append(current_line)

    return lines

# Create certificates
def generate_certificates(data, template_path, output_dir):
    # Open the template
    template = Image.open(template_path)
    if not os.path.exists(output_dir):
        os.makedirs(output_dir)
    
    # Set font and text positions
    font_path = "arial.ttf"  # Path to your font file
    name_font = ImageFont.truetype(font_path, size=50)
    title_font = ImageFont.truetype(font_path, size=40)

    # Predefined positions (adjust based on your template)
    name_y_position = 500  # Y-coordinate for Name
    title_y_start = 620  # Starting Y-coordinate for Title Name
    canvas_width = template.width  # Get template width for centering
    max_title_width = int(canvas_width * 0.8)  # Allow title to use 80% of the canvas width

    for i, (name, title) in enumerate(data, start=1):
        cert = template.copy()
        draw = ImageDraw.Draw(cert)
        
        # Center the Name horizontally
        name_bbox = draw.textbbox((0, 0), name, font=name_font)
        name_width = name_bbox[2] - name_bbox[0]
        name_x_position = (canvas_width - name_width) // 2
        draw.text((name_x_position, name_y_position), name, fill="black", font=name_font)

        # Wrap Title Name and center each line
        wrapped_lines = wrap_text(title, title_font, max_title_width, draw)
        line_spacing = 10  # Spacing between lines
        y_position = title_y_start

        for line in wrapped_lines:
            title_bbox = draw.textbbox((0, 0), line, font=title_font)
            title_width = title_bbox[2] - title_bbox[0]
            title_x_position = (canvas_width - title_width) // 2
            draw.text((title_x_position, y_position), line, fill="black", font=title_font)
            y_position += title_bbox[3] - title_bbox[1] + line_spacing

        # Save certificate
        output_path = os.path.join(output_dir, f"Certificate_{i}_{name}.png")
        cert.save(output_path)
        print(f"Created: {output_path}")

# Main function
def main():
    print("Reading Excel file...")
    data = load_excel_data(excel_file)
    
    print("Generating certificates...")
    generate_certificates(data, template_image, output_folder)
    print("All certificates are ready!")

# Run the program
if __name__ == "__main__":
    main()
